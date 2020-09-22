import xlwt
import matplotlib.pyplot as plt
import cv2
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill


class tools(object):
    '''
    1. 只画 img
    2. 画 img 和 原图
    3. 画 img ，原图 和 标记图
    '''

    # def imgToExcel(self, flag, saveFile,binImg, origin = [], markArr = [], markArr2 = []):
    #     pattern = xlwt.Pattern()
    #     pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    #     pattern.pattern_fore_colour = 2  # 绿
    #     style = xlwt.XFStyle()  # 噪声
    #     style.pattern = pattern
    #     book = xlwt.Workbook(encoding='utf-8', style_compression=0)
    #     sheet = book.add_sheet('m', cell_overwrite_ok=True)
    #     count, tj = 0, 0
    #     h, w = binImg.shape
    #     for j in range(w):
    #         if tj > 250:
    #             tj = 0
    #             name_sheet = 'm' + str(int(j / 255));
    #             sheet = book.add_sheet(name_sheet, cell_overwrite_ok=True)
    #         for i in range(h):
    #             if binImg[i,j]:
    #                 # sheet.write(i, tj, int(markArr[i, j]), style)
    #                 # sheet.write(i, tj, int(img[i, j]))
    #                 if flag == 0:
    #                     sheet.write(i, tj, str(origin[i, j]))
    #                 elif flag == 3:
    #                     sheet.write(i, tj, str(int(origin[i ,j]))+'/'+str(int(markArr[i,j])), style)
    #                 elif flag == 2:
    #                     # 标记噪声图 专用
    #                     sheet.write(i, tj, str(int(origin[i,j]))+'/'+str(int(binImg[i,j])), style)
    #                 elif flag == 1:
    #                     sheet.write(i, tj, str(origin[i,j]), style)
    #                 elif flag == 4:
    #                     sheet.write(i, tj , str(int(origin[i,j])) + '/' + str(int(markArr[i, j])) + '/' + str(int(markArr2[i,j])), style)
    #
    #
    #             else:
    #                 if flag == 0:
    #                     sheet.write(i, tj, str(origin[i, j]))
    #                 elif flag == 3:
    #                     sheet.write(i, tj, str(int(origin[i, j])) + '/' + str(int(markArr[i, j])))
    #                 elif flag == 2:
    #                     # 标记噪声图 专用
    #                     sheet.write(i, tj, str(int(origin[i,j]))+'/'+str(int(binImg[i,j])))
    #                 elif flag == 1:
    #                     sheet.write(i, tj, str(origin[i,j]))
    #                 elif flag == 4:
    #                     sheet.write(i, tj , str(int(origin[i,j])) + '/' + str(int(markArr[i, j])) + '/' + str(int(markArr2[i,j])))
    #
    #
    #         tj += 1
    #     book.save(saveFile + '.xls')
    #     print("imgToexcel OK")

    # '''
    # 4-1  1水平竖直一个颜色
    # 正斜线 一个颜色
    # 反斜线 一个颜色
    # 同时正反一个颜色
    # '''
    # def imgToExcel_4_1(self, saveFile,binImg, origin = []):
    #     # 水平
    #     pattern = xlwt.Pattern()
    #     pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    #     pattern.pattern_fore_colour = 3  # 绿
    #     style = xlwt.XFStyle()  #
    #     style.pattern = pattern
    #     # 竖直
    #     pattern1 = xlwt.Pattern()
    #     pattern1.pattern = xlwt.Pattern.SOLID_PATTERN
    #     pattern1.pattern_fore_colour = 3  # 绿
    #     style1 = xlwt.XFStyle()  #
    #     style1.pattern = pattern1
    #
    #     # 正对角线
    #     pattern2 = xlwt.Pattern()
    #     pattern2.pattern = xlwt.Pattern.SOLID_PATTERN
    #     pattern2.pattern_fore_colour = 4  # 绿
    #     style2 = xlwt.XFStyle()  #
    #     style2.pattern = pattern2
    #
    #     # 反对角线
    #     pattern3 = xlwt.Pattern()
    #     pattern3.pattern = xlwt.Pattern.SOLID_PATTERN
    #     pattern3.pattern_fore_colour = 5  # 绿
    #     style3 = xlwt.XFStyle()  #
    #     style3.pattern = pattern3
    #
    #     # 正反对角线
    #     pattern4 = xlwt.Pattern()
    #     pattern4.pattern = xlwt.Pattern.SOLID_PATTERN
    #     pattern4.pattern_fore_colour = 6  # 绿
    #     style4 = xlwt.XFStyle()  #
    #     style4.pattern = pattern4
    #
    #     book = xlwt.Workbook(encoding='utf-8', style_compression=0)
    #     sheet = book.add_sheet('m', cell_overwrite_ok=True)
    #     h, w = binImg.shape
    #     # 原始数组元素
    #     isFirst = False
    #     tj = 0
    #     for j in range(0,w, 2):
    #         if tj > 250:
    #             tj = 0
    #             if isFirst==False:
    #                 name_sheet = 'm' + str(int(j / 255));
    #                 sheet = book.add_sheet(name_sheet, cell_overwrite_ok=True)
    #                 isFirst = True
    #         for i in range(0, h, 2):
    #             sheet.write(i, tj, str(origin[i, j]))
    #         tj += 2
    #     # 水平 从第1列开始，行从第0行
    #     tj = 1
    #     for j in range(1,w, 2):
    #         if tj > 250:
    #             tj = 1
    #             if isFirst == False:
    #                 name_sheet = 'm' + str(int(j / 255));
    #                 sheet = book.add_sheet(name_sheet, cell_overwrite_ok=True)
    #                 isFirst = True
    #         for i in range(0, h, 2):
    #             if binImg[i,j] > 0:
    #                 sheet.write(i, tj, str(origin[i, j]), style)
    #             else:
    #                 sheet.write(i, tj, str(origin[i, j]))
    #         tj += 2
    #
    #
    #     # 竖直 从第0列开始，行从第1行
    #     tj = 0
    #     for j in range(0,w, 2):
    #         if tj > 250:
    #
    #             tj = 0
    #             if isFirst==False:
    #                 name_sheet = 'm' + str(int(j / 255));
    #                 sheet = book.add_sheet(name_sheet, cell_overwrite_ok=True)
    #                 isFirst= True
    #         for i in range(1, h, 2):
    #             if binImg[i,j] > 0:
    #                 sheet.write(i, tj, str(origin[i, j]), style1)
    #             else:
    #                 sheet.write(i, tj, str(origin[i, j]))
    #         tj += 2
    #
    #
    #     # 对角 从第1列开始，行从第1行
    #     tj = 1
    #     for j in range(1,w, 2):
    #         if tj > 250:
    #             tj = 1
    #             if isFirst==False:
    #                 name_sheet = 'm' + str(int(j / 255));
    #                 sheet = book.add_sheet(name_sheet, cell_overwrite_ok=True)
    #                 isFirst=True
    #         for i in range(1, h, 2):
    #             if binImg[i,j] > 0:
    #                 if binImg[i,j] == 3:
    #                     sheet.write(i, tj, str(origin[i, j]), style2)
    #                 elif binImg[i,j] == 4:
    #                     sheet.write(i, tj, str(origin[i, j]), style3)
    #                 else:
    #                     sheet.write(i, tj, str(origin[i, j]), style4)
    #
    #             else:
    #                 sheet.write(i, tj, str(origin[i, j]))
    #         tj += 2
    #     book.save(saveFile + '.xls')
    #     print("imgToexcel OK")
    # 将 diff 不同的地方用另外一种颜色显示出来
    # img 原图 diff 单独染色的图
    # 原图的像素值-标记， 新图的像素值-标记， 添加的噪声点
    # def twoImgToExcel(self,  saveFile,originPix, originMark, newPix, newMark, noiseCoordinate):
    #     pattern = xlwt.Pattern()
    #     pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    #     pattern.pattern_fore_colour = 2  # 绿
    #     style = xlwt.XFStyle()  #
    #     style.pattern = pattern
    #
    #     pattern1 = xlwt.Pattern()
    #     pattern1.pattern = xlwt.Pattern.SOLID_PATTERN
    #     pattern1.pattern_fore_colour = 3  # 绿
    #     style1 = xlwt.XFStyle()  #
    #     style1.pattern = pattern1
    #
    #     book = xlwt.Workbook(encoding='utf-8', style_compression=0)
    #     sheet = book.add_sheet('m', cell_overwrite_ok=True)
    #     count, tj = 0, 0
    #     h, w = originPix.shape
    #     successNum = 0
    #     for j in range(w):
    #         if tj > 250:
    #             tj = 0
    #             name_sheet = 'm' + str(int(j / 255));
    #             sheet = book.add_sheet(name_sheet, cell_overwrite_ok=True)
    #         for i in range(h):
    #             s = str(originPix[i,j])+'-'+str(originMark[i, j])+'/'+str(newPix[i, j])+'-'+str(newMark[i, j])
    #             if newMark[i,j]:
    #                 sheet.write(i, tj, s,style)
    #             else:
    #                 sheet.write(i, tj, s)
    #             if (i, j) in noiseCoordinate:
    #                 sheet.write(i, tj, s, style1)
    #                 successNum += 1
    #
    #
    #
    #         tj += 1
    #     book.save(saveFile + '.xls')
    #     print('successNum={}'.format(successNum))
    #     print("imgToexcel OK")
    # 画直方图
    def drawHist(self, lists, xName="", yName=""):
        fig = plt.figure()
        ax = fig.add_subplot(1, 1, 1)
        right = max(lists)
        print(xName + 'maxV=', right)
        plt.hist(lists, right, [0, right]);
        lens = int(right / 20)
        # lens = 1
        x = range(0, right, lens)
        plt.xticks(x)
        ax.set_xlabel(xName)
        ax.set_ylabel(yName)
        plt.show()

    # 针对叉边方法画图
    def drawBinPicByChabian(self, msg, binImg):
        h, w = binImg.shape
        img = binImg.copy()
        for i in range(h):
            for j in range(w):
                if 0 < binImg[i, j] < 10:
                    img[i, j] = 255
                else:
                    img[i, j] = 0

        cv2.imshow(msg, img)

    # msg -> 图像名字
    # binImg -> 二值图像
    def drawBinPic(self, msg, binImg):
        h, w = binImg.shape
        img = binImg.copy()
        for i in range(h):
            for j in range(w):
                if binImg[i, j] > 0:
                    img[i, j] = 255
                else:
                    img[i, j] = 0
        cv2.imshow(msg, img)

    # msg -> 名字
    # img -> 图像
    def imgToCSV(self, msg, img):
        markPd = pd.DataFrame(img)
        markPd.to_csv(msg + '.csv')

    def markPicToBinPic8(self, markPic):
        h, w = markPic.shape
        binPic = np.zeros(markPic.shape, np.uint8)

        def check(dx, dy):
            if dx < 0 or dx >= h or dy < 0 or dy >= w:
                return True
            return False

        for i in range(h):
            for j in range(w):
                total, same = 0, 0
                flag = False
                for ii in (-1, 2):
                    for jj in (-1, 2):
                        dx, dy = i + ii, j + jj
                        if check(dx, dy):
                            continue
                        total += 1
                        if markPic[dx, dy] == markPic[i, j]:
                            same += 1
                        else:
                            flag = True
                            break;
                    if flag:
                        break;
                if total > same:
                    binPic[i, j] = 255
        return binPic

    def markPicToBinPic4(self, markPic):
        h, w = markPic.shape
        binPic = np.zeros(markPic.shape, np.uint8)

        def check(dx, dy):
            if dx < 0 or dx >= h or dy < 0 or dy >= w:
                return True
            return False

        dir = [-1, 0, 1, 0, -1]
        for i in range(h):
            for j in range(w):
                total, same = 0, 0
                for k in range(4):
                    dx, dy = i + dir[k], j + dir[k + 1]
                    if check(dx, dy):
                        continue
                    total += 1
                    if markPic[dx, dy] == markPic[i, j]:
                        same += 1
                    else:
                        break;
                if total > same:
                    binPic[i, j] = 255
        return binPic

    def imgToExcel_bsm(self, saveFile, mark, origin=[], ourimg=[], ground=[]):
        # 水平

        wb = Workbook()
        sheet = wb.active
        h, w = mark.shape

        fill1 = PatternFill("solid", fgColor="76EE00")  # 绿色
        fill2 = PatternFill("solid", fgColor="EE7AE9")  # 粉红
        fill3 = PatternFill("solid", fgColor="F0E68C")  # 淡黄色
        # 原始数组元素
        for i in range(h):
            for j in range(w):
                ii, jj = i // 2, j // 2
                # if mark[i,j]==5:
                #     sheet.write(i, tj, str(origin[ii, jj]))
                # elif mark[i,j]==6:
                #     sheet.write(i, tj, str(origin[ii, jj]))
                # elif mark[i, j] == 7:
                #     sheet.write(i, tj, str(origin[ii, jj]))
                # else:
                #     sheet.write(i, tj, str(origin[ii, jj]))
                ki, kj = i + 1, j + 1
                sheet.cell(ki, kj, str(mark[i, j]))
                # ourimg
                if i % 2 == 0 and j % 2 == 0:
                    sheet.cell(ki, kj, str(origin[ii, jj]))
                    if ourimg[ii, jj] > 0:
                        sheet.cell(ki, kj, str(origin[ii, jj]))  # 绿色
                        sheet.cell(ki, kj).fill = fill1
                    if ground[ii, jj] > 0:
                        sheet.cell(ki, kj, str(origin[ii, jj]))  # 粉红
                        sheet.cell(ki, kj).fill = fill2
                    if ourimg[ii, jj] > 0 and ground[ii, jj] > 0:
                        sheet.cell(ki, kj, str(origin[ii, jj]))  # 淡黄色
                        sheet.cell(ki, kj).fill = fill3

        wb.save(saveFile + '.xlsx')
        print("imgToexcel OK")

    '''
    origin: 原始mat
    ourimg: 边缘二值mat
    ground: groundTruth mat
    '''

    def imgToExcel_edge_groundTruth(self, saveFile, origin=[], ourimg=[], ground=[]):
        # 水平
        wb = Workbook()
        sheet = wb.active

        fill1 = PatternFill("solid", fgColor="76EE00")  # 绿色
        fill2 = PatternFill("solid", fgColor="EE7AE9")  # 粉红
        fill3 = PatternFill("solid", fgColor="F0E68C")  # 淡黄色

        h, w = origin.shape
        # 原始数组元素
        for i in range(h):
            for j in range(w):
                # ourimg
                ii, jj = i + 1, j + 1
                sheet.cell(ii, jj, str(origin[i, j]))
                if ourimg[i, j] > 0:
                    sheet.cell(ii, jj, str(origin[i, j]))  # 绿色
                    sheet.cell(ii, jj).fill = fill1
                if ground[i, j] > 0:
                    sheet.cell(ii, jj, str(origin[i, j]))  # 粉红
                    sheet.cell(ii, jj).fill = fill2

                if ourimg[i, j] > 0 and ground[i, j] > 0:
                    sheet.cell(ii, jj, str(origin[i, j]))  # 淡黄色
                    sheet.cell(ii, jj).fill = fill3
        wb.save(saveFile + '.xlsx')
        print("imgToexcel OK")

    def grayAndBin(self, saveFile, origin=[], ground=[]):
        # 水平
        wb = Workbook()
        sheet = wb.active

        fill1 = PatternFill("solid", fgColor="76EE00")  # 绿色
        # fill2 = PatternFill("solid", fgColor="EE7AE9")  # 粉红
        # fill3 = PatternFill("solid", fgColor="F0E68C")  # 淡黄色

        h, w = origin.shape
        # 原始数组元素
        for i in range(h):
            for j in range(w):
                # ourimg
                ii, jj = i + 1, j + 1
                sheet.cell(ii, jj, str(origin[i, j]))
                if ground[i, j] > 0:
                    sheet.cell(ii, jj, str(origin[i, j]))
                    sheet.cell(ii, jj).fill = fill1

        wb.save(saveFile + '.xlsx')
        print(saveFile + "OK")

    '''
    将单一mat 大于0 的位置画出来
    '''

    def singleAndGreaterThanZero(self, saveFile, gray, bin):
        # 水平
        wb = Workbook()
        sheet = wb.active

        fill1 = PatternFill("solid", fgColor="76EE00")  # 绿色
        # fill2 = PatternFill("solid", fgColor="EE7AE9")  # 粉红
        # fill3 = PatternFill("solid", fgColor="F0E68C")  # 淡黄色

        h, w = bin.shape
        # 原始数组元素
        for i in range(h):
            for j in range(w):
                # ourimg
                ii, jj = i + 1, j + 1
                # sheet.cell(ii, jj, str(gray[i,j]) + '/' +str(bin[i, j]))
                sheet.cell(ii, jj, str(gray[i, j]))
                if bin[i, j] > 0:
                    # sheet.cell(ii, jj, str(gray[i,j]) + '/' +str(bin[i, j]))
                    sheet.cell(ii, jj).fill = fill1

        wb.save(saveFile + '.xlsx')
        print(saveFile + "OK")

    def gray_dmat_mark_sho_to_excel(self, saveFile, gray, dmat, bin, sho, ground_mat):
        wb = Workbook()
        sheet = wb.active

        fill1 = PatternFill("solid", fgColor="76EE00")  # 绿色
        fill2 = PatternFill("solid", fgColor="EE7AE9")  # 粉红 groundTruth
        fill3 = PatternFill("solid", fgColor="F0E68C")  # 淡黄色 我们的和groundtruth共有的

        h, w = bin.shape
        # 原始数组元素
        for i in range(h):
            for j in range(w):
                # ourimg
                ii, jj = i + 1, j + 1
                sheet.cell(ii, jj, str(gray[i, j]) + '/' + str(dmat[i, j]) + '/' + str(sho[i, j]))

                # 如果是共有的，就画fill3
                if bin[i, j] > 0 and ground_mat[i, j] > 0:
                    sheet.cell(ii, jj).fill = fill3
                else:
                    # 我们的边缘点，画绿色
                    if bin[i, j] > 0:
                        # sheet.cell(ii, jj, str(gray[i,j]) + '/' + str(dmat[i, j]) +  '/' + str(sho[i,j]))
                        sheet.cell(ii, jj).fill = fill1

                    # groundTruth, 画红色
                    if ground_mat[i, j] > 0:
                        sheet.cell(ii, jj).fill = fill2

        wb.save(saveFile + '.xlsx')
        print(saveFile + "OK")

    ''''
    5-24
    gray, 区分度，二值边缘， groundTruth
    '''

    def gray_dmat_mark_to_excel(self, saveFile, gray, dmat, bin, ground_mat):
        wb = Workbook()
        sheet = wb.active

        fill1 = PatternFill("solid", fgColor="76EE00")  # 绿色
        fill2 = PatternFill("solid", fgColor="EE7AE9")  # 粉红 groundTruth
        fill3 = PatternFill("solid", fgColor="F0E68C")  # 淡黄色 我们的和groundtruth共有的

        h, w = bin.shape
        # 原始数组元素
        for i in range(h):
            for j in range(w):
                # ourimg
                ii, jj = i + 1, j + 1
                sheet.cell(ii, jj, str(gray[i, j]) + '/' + str(dmat[i, j]))

                # 如果是共有的，就画fill3
                if bin[i, j] > 0 and ground_mat[i, j] > 0:
                    sheet.cell(ii, jj).fill = fill3
                else:
                    # 我们的边缘点，画绿色
                    if bin[i, j] > 0:
                        # sheet.cell(ii, jj, str(gray[i,j]) + '/' + str(dmat[i, j]) +  '/' + str(sho[i,j]))
                        sheet.cell(ii, jj).fill = fill1

                    # groundTruth, 画红色
                    if ground_mat[i, j] > 0:
                        sheet.cell(ii, jj).fill = fill2

        wb.save(saveFile + '.xlsx')
        print(saveFile + "OK")

    def gray_dmat_to_excel1(self, saveFile, gray, dmat):
        # 水平
        wb = Workbook()
        sheet = wb.active

        fill1 = PatternFill("solid", fgColor="76EE00")  # 绿色
        # fill2 = PatternFill("solid", fgColor="EE7AE9")  # 粉红
        # fill3 = PatternFill("solid", fgColor="F0E68C")  # 淡黄色

        h, w = gray.shape
        # 原始数组元素
        for i in range(h):
            for j in range(w):
                # ourimg
                ii, jj = i + 1, j + 1
                sheet.cell(ii, jj, str(gray[i, j]) + '/' + str(dmat[i, j]))
                if dmat[i, j] > 10:
                    sheet.cell(ii, jj, str(gray[i, j]) + '/' + str(dmat[i, j]))
                    sheet.cell(ii, jj).fill = fill1

        wb.save(saveFile + '.xlsx')
        print(saveFile + "OK")

    def min_max_inter_max_to_excel(self, saveFile, mat, ret_mat, inter_mat, ground_mat=[]):
        wb = Workbook()
        sheet = wb.active

        fill1 = PatternFill("solid", fgColor="76EE00")  # 绿色 应用规则方法前
        fill2 = PatternFill("solid", fgColor="EE7AE9")  # 粉红 应用规则方法后
        fill3 = PatternFill("solid", fgColor="F0E68C")  # 淡黄色 两种方法共同的
        fill4 = PatternFill("solid", fgColor="6A5ACD")  # 紫色 groundTruth
        fill5 = PatternFill("solid", fgColor="63B8FF")  # 蓝色 ground与方法前同一种颜色
        fill6 = PatternFill("solid", fgColor="76EE00")  # 灰色 ground 与方法后同一种颜色
        fill7 = PatternFill("solid", fgColor="CD2626")  # 红色 三种共有的颜色

        h, w = mat.shape
        # 原始数组元素
        for i in range(h):
            for j in range(w):
                # ourimg
                ii, jj = i + 1, j + 1
                sheet.cell(ii, jj, str(mat[i, j]) + '/' + str(inter_mat[i, j]) + '/' + str(ret_mat[i, j]))

                if ret_mat[i, j] == 1:
                    sheet.cell(ii, jj).fill = fill1
                elif ret_mat[i, j] == 10:
                    sheet.cell(ii, jj).fill = fill2
                elif ret_mat[i, j] == 100:
                    sheet.cell(ii, jj).fill = fill3
                elif ret_mat[i, j] == 110:
                    sheet.cell(ii, jj).fill = fill4
                elif ret_mat[i, j] == 101:
                    sheet.cell(ii, jj).fill = fill5

        wb.save(saveFile + '.xlsx')
        print(saveFile + "OK")

    '''
    修复前后是由什么引起的
    
    '''

    def after_repair_show_diff_to_excel(self, saveFile, mat, ret_mat, d_mat_before, repair_mat, d_mat_after, bin,
                                        ground_mat=[]):
        wb = Workbook()
        sheet = wb.active

        # fill1 = PatternFill("solid", fgColor="76EE00")  # 绿色 应用规则方法前
        # fill2 = PatternFill("solid", fgColor="EE7AE9")  # 粉红 应用规则方法后
        # fill3 = PatternFill("solid", fgColor="F0E68C")  # 淡黄色 两种方法共同的
        # fill4 = PatternFill("solid", fgColor="6A5ACD")  # 紫色 groundTruth
        # fill5 = PatternFill("solid", fgColor="63B8FF")  # 蓝色 ground与方法前同一种颜色
        # fill6 = PatternFill("solid", fgColor="76EE00")  # 灰色 ground 与方法后同一种颜色
        # fill7 = PatternFill("solid", fgColor="CD2626")  # 红色 三种共有的颜色
        fill1 = PatternFill("solid", fgColor="76EE00")  # 绿 big edge
        fill2 = PatternFill("solid", fgColor="EE7AE9")  # 红 small edge
        fill3 = PatternFill("solid", fgColor="F0E68C")  # 淡黄色 我们的和groundtruth共有的
        h, w = mat.shape
        # 原始数组元素
        for i in range(h):
            for j in range(w):
                # ourimg
                ii, jj = i + 1, j + 1
                if ret_mat[i, j] == 0:
                    sheet.cell(ii, jj, str(mat[i, j]) + '/' + str(d_mat_after[i, j]))
                else:
                    sheet.cell(ii, jj,
                               str(mat[i, j]) + '/' + str(d_mat_before[i, j]) + '/' + str(ret_mat[i, j]) + '/' + str(
                                   repair_mat[i, j]) + '/' + str(d_mat_after[i, j]))

                # if ret_mat[i, j] == 1:
                #     sheet.cell(ii, jj).fill = fill1
                # elif ret_mat[i, j] == 10:
                #     sheet.cell(ii, jj).fill = fill2
                # elif ret_mat[i, j] == 100:
                #     sheet.cell(ii, jj).fill = fill3
                # elif ret_mat[i, j] == 110:
                #     sheet.cell(ii, jj).fill = fill4
                # elif ret_mat[i, j] == 101:
                #     sheet.cell(ii, jj).fill = fill5
                if bin[i, j] >= 10:
                    sheet.cell(ii, jj).fill = fill1
                elif 1 <= bin[i, j] < 10:
                    sheet.cell(ii, jj).fill = fill2
                elif bin[i, j] == 110:
                    sheet.cell(ii, jj).fill = fill3

        wb.save(saveFile + '.xlsx')
        print(saveFile + "OK")

    def after_repair_min_max_inter_max_to_excel(self, saveFile, mat, ret_mat, inter_mat, repair_mat, ground_mat=[]):
        wb = Workbook()
        sheet = wb.active

        fill1 = PatternFill("solid", fgColor="76EE00")  # 绿色 应用规则方法前
        fill2 = PatternFill("solid", fgColor="EE7AE9")  # 粉红 应用规则方法后
        fill3 = PatternFill("solid", fgColor="F0E68C")  # 淡黄色 两种方法共同的
        fill4 = PatternFill("solid", fgColor="6A5ACD")  # 紫色 groundTruth
        fill5 = PatternFill("solid", fgColor="63B8FF")  # 蓝色 ground与方法前同一种颜色
        fill6 = PatternFill("solid", fgColor="76EE00")  # 灰色 ground 与方法后同一种颜色
        fill7 = PatternFill("solid", fgColor="CD2626")  # 红色 三种共有的颜色

        h, w = mat.shape
        # 原始数组元素
        for i in range(h):
            for j in range(w):
                # ourimg
                ii, jj = i + 1, j + 1
                if ret_mat[i, j] == 0:
                    sheet.cell(ii, jj, str(mat[i, j]) + '/' + str(inter_mat[i, j]) + '/' + str(ret_mat[i, j]))
                else:
                    sheet.cell(ii, jj,
                               str(mat[i, j]) + '/' + str(inter_mat[i, j]) + '/' + str(ret_mat[i, j]) + '/' + str(
                                   repair_mat[i, j]))

                if ret_mat[i, j] == 1:
                    sheet.cell(ii, jj).fill = fill1
                elif ret_mat[i, j] == 10:
                    sheet.cell(ii, jj).fill = fill2
                elif ret_mat[i, j] == 100:
                    sheet.cell(ii, jj).fill = fill3
                elif ret_mat[i, j] == 110:
                    sheet.cell(ii, jj).fill = fill4
                elif ret_mat[i, j] == 101:
                    sheet.cell(ii, jj).fill = fill5

        wb.save(saveFile + '.xlsx')
        print(saveFile + "OK")

    def gray_dmat_sho_to_excel(self, saveFile, gray, d_mat, e_mat_before, e_mat_after, sho_mat, ground_mat):
        wb = Workbook()
        sheet = wb.active

        fill1 = PatternFill("solid", fgColor="76EE00")  # 绿色 应用规则方法前
        fill2 = PatternFill("solid", fgColor="EE7AE9")  # 粉红 应用规则方法后
        fill3 = PatternFill("solid", fgColor="F0E68C")  # 淡黄色 两种方法共同的
        fill4 = PatternFill("solid", fgColor="6A5ACD")  # 紫色 groundTruth
        fill5 = PatternFill("solid", fgColor="63B8FF")  # 蓝色 ground与方法前同一种颜色
        fill6 = PatternFill("solid", fgColor="76EE00")  # 灰色 ground 与方法后同一种颜色
        fill7 = PatternFill("solid", fgColor="CD2626")  # 红色 三种共有的颜色

        h, w = gray.shape
        # 原始数组元素
        for i in range(h):
            for j in range(w):
                # ourimg
                ii, jj = i + 1, j + 1
                sheet.cell(ii, jj, str(gray[i, j]) + '/' + str(d_mat[i, j]) + '/' + str(sho_mat[i, j]))

                if e_mat_before[i, j] > 0 and e_mat_after[i, j] > 0 and ground_mat[i, j] > 0:
                    sheet.cell(ii, jj).fill = fill7

                elif e_mat_before[i, j] > 0 and e_mat_after[i, j] > 0:
                    sheet.cell(ii, jj).fill = fill3

                elif e_mat_before[i, j] > 0 and ground_mat[i, j] > 0:
                    sheet.cell(ii, jj).fill = fill5

                elif e_mat_after[i, j] > 0 and ground_mat[i, j] > 0:
                    sheet.cell(ii, jj).fill = fill6
                else:
                    if e_mat_before[i, j] > 0:
                        sheet.cell(ii, jj).fill = fill1
                    if e_mat_after[i, j] > 0:
                        sheet.cell(ii, jj).fill = fill2
                    if ground_mat[i, j] > 0:
                        sheet.cell(ii, jj).fill = fill4

                # # 如果是共有的，就画fill3
                # if bin[i, j] > 0 and ground_mat[i, j] > 0:
                #     sheet.cell(ii, jj).fill = fill3
                # else:
                #     # 我们的边缘点，画绿色
                #     if bin[i, j] > 0:
                #         # sheet.cell(ii, jj, str(gray[i,j]) + '/' + str(dmat[i, j]) +  '/' + str(sho[i,j]))
                #         sheet.cell(ii, jj).fill = fill1
                #
                #     # groundTruth, 画红色
                #     if ground_mat[i, j] > 0:
                #         sheet.cell(ii, jj).fill = fill2

        wb.save(saveFile + '.xlsx')
        print(saveFile + "OK")

    def gray_dmat_big_small_to_excel(self, saveFile, gray, dmat, bin, inter_mat, ground_mat=[]):
        wb = Workbook()
        sheet = wb.active

        fill1 = PatternFill("solid", fgColor="76EE00")  # 绿 big edge
        fill2 = PatternFill("solid", fgColor="EE7AE9")  # 红 small edge
        fill3 = PatternFill("solid", fgColor="F0E68C")  # 淡黄色 我们的和groundtruth共有的

        h, w = gray.shape
        # 原始数组元素
        for i in range(h):
            for j in range(w):
                # ourimg
                ii, jj = i + 1, j + 1
                sheet.cell(ii, jj, str(gray[i, j]) + '/' + str(dmat[i, j]))

                # 如果是共有的，就画fill3
                if bin[i, j] >= 10:
                    sheet.cell(ii, jj).fill = fill1
                elif 1 <= bin[i, j] < 10:
                    sheet.cell(ii, jj).fill = fill2
                elif bin[i, j] == 110:
                    sheet.cell(ii, jj).fill = fill3

        wb.save(saveFile + '.xlsx')
        print(saveFile + "OK")

    '''
    修复后多余的点用黄色和粉红表示
    '''

    def repair_before_and_after_to_excel(self, excel_name, ori_mat, before_bin, d_mat_before, repaired_pix_mat,
                                         d_mat_after, after_bin, ground_mat=[]):
        wb = Workbook()
        sheet = wb.active

        fill1 = PatternFill("solid", fgColor="76EE00")  # 红色 修复前小边
        fill2 = PatternFill("solid", fgColor="EE7AE9")  # 绿色 修复前大边
        fill3 = PatternFill("solid", fgColor="F0E68C")  # 黄色 修复后小边
        fill4 = PatternFill("solid", fgColor="6A5ACD")  # 粉红 修复后大边
        fill5 = PatternFill("solid", fgColor="63B8FF")  # 蓝色 ground与方法前同一种颜色
        fill6 = PatternFill("solid", fgColor="76EE00")  # 灰色 ground 与方法后同一种颜色
        fill7 = PatternFill("solid", fgColor="CD2626")  # 红色 三种共有的颜色

        h, w = ori_mat.shape
        # 原始数组元素
        for i in range(h):
            for j in range(w):
                # ourimg
                ii, jj = i + 1, j + 1
                if before_bin[i, j] > 0:
                    info = str(ori_mat[i, j]) + '/' + str(d_mat_before[i, j]) + '/' + \
                           str(repaired_pix_mat[i, j]) + '/' + str(d_mat_after[i, j]) + '/' + \
                           str(before_bin[i, j])

                elif after_bin[i, j] > 0:
                    info = str(ori_mat[i, j]) + '/' + str(d_mat_before[i, j]) + '/' + \
                           str(repaired_pix_mat[i, j]) + '/' + str(d_mat_after[i, j]) + '/' + \
                           str(after_bin[i, j])
                else:
                    info = str(ori_mat[i, j]) + '/' + str(d_mat_before[i, j])

                sheet.cell(ii, jj, info)

                if after_bin[i, j] >= 10:
                    sheet.cell(ii, jj).fill = fill3
                elif 1 <= after_bin[i, j] < 10:
                    sheet.cell(ii, jj).fill = fill4

                if before_bin[i, j] >= 10:
                    sheet.cell(ii, jj).fill = fill1
                elif 1 <= before_bin[i, j] < 10:
                    sheet.cell(ii, jj).fill = fill2

        wb.save(excel_name + '.xlsx')
        print(excel_name + "\tOK")
